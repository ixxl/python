import xlrd
import xlwt
import paramiko
import os
import time
import re
import openpyxl
from openpyxl import Workbook

CSheet = xlrd.open_workbook(r'C:\Users\Administrator\Desktop\IPlist.xls')
data = CSheet.sheets()[0]
nrow = data.nrows

FailedList = []

SuccessList = []


for i in range(1, nrow):
    ip = data.row_values(i)[0]

    try:
        transport = paramiko.Transport((ip, 22))  #进入后台
        transport.connect(username='root', password='dhERIS@2018*#')

        ssh = paramiko.SSHClient()
        # 自动添加策略，保存服务器的主机名和秘钥信息，不添加，不在本地hnow_hosts文件中的记录将无法连接
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh._transport = transport

        print("%s 连接成功" % (ip))

        # 查看磁盘情况
        stdin, stdout, stderror = ssh.exec_command("smartctl -l error /dev/sda | grep Logged")  
        state = stdout.readlines()[0]
        print(state)
        
        state = state.rstrip() #去除字符串
 #判断状态       
        if state == "No Errors Logged" : 
            SuccessList.append(ip)
        else :
            FailedList.append(ip)

        

    except Exception as reason:
        print(reason)
        FailedList.append(ip)

# 创建表格
wb = Workbook()
ws = wb.create_sheet('sheet', 0)


for i in range(len(SuccessList)):  
    ws.cell(row=i+2,column=1).value = SuccessList[i]
for j in range(len(FailedList)):  
    ws.cell(row=j+2,column=2).value = FailedList[j]
    
wb.save(r'C:\Users\Administrator\Desktop\硬盘.xlsx')  #表格保存路径
print('文件 导出成功！')

print("连接失败工控机：")
for j in FailedList:
    print(j)
